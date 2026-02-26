"""Simple FastAPI application for testing."""

import asyncio
import smtplib
import time
from contextlib import asynccontextmanager
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
from jose import jwt
from fastapi import FastAPI, HTTPException, Body, Header, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel
import random
import warnings
import pandas as pd
from pathlib import Path

from app.config import get_settings
from app import redis_cache
from app import access_allowlist
from app import compliance_db


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Handle startup/shutdown. Warm Excel cache so first request is fast; suppress CancelledError on reload."""
    # Warm in-memory Excel cache in a thread so first compliance request doesn't block on read
    try:
        _data_dir = Path(__file__).resolve().parent / "data"
        _compliance_file = _data_dir / "Compliance_Register - Copy.xlsx"
        if _data_dir.exists() and _compliance_file.exists():
            await asyncio.to_thread(read_all_sheets_from_excel, _compliance_file)
    except Exception:
        pass
    # Compliance SQLite: import Excel into DB once so reads are fast (no Excel on every request)
    try:
        from app.config import get_settings as _gs
        _s = _gs()
        if _s.use_compliance_sqlite and COMPLIANCE_FILE.exists():
            await asyncio.to_thread(compliance_db.import_from_excel, read_all_sheets_from_excel, COMPLIANCE_FILE)
            # Warm Redis with default initial response (limit=50, no filters) so first request is a cache hit
            if getattr(_s, "use_redis_cache", False) and compliance_db.has_data():
                try:
                    depts = compliance_db.get_departments()
                    statuses = compliance_db.get_statuses()
                    recs, total = compliance_db.get_records(50, 0, None, None)
                    for r in recs:
                        for k, v in list(r.items()):
                            if v is not None and isinstance(v, float) and (v != v or v == 1e999):
                                r[k] = None
                    out = {"success": True, "departments": depts, "statuses": statuses, "records": {"data": recs, "total": total, "limit": 50, "offset": 0}}
                    redis_cache.cache_set("compliance:initial:50:0::", out, getattr(_s, "redis_cache_ttl", 300))
                except Exception:
                    pass
    except Exception:
        pass
    # Ensure access allowlist DB and table exist; sync allowed usernames from file (no names in code)
    try:
        access_allowlist.init_db()
        await asyncio.to_thread(access_allowlist.sync_from_file)
    except Exception:
        pass
    try:
        yield
    except asyncio.CancelledError:
        pass  # normal when uvicorn --reload restarts the worker


# Create FastAPI application
app = FastAPI(
    title="Fraud Detection Backend (Test Mode)",
    description="Simple test API for fraud detection system",
    version="1.0.0-test",
    docs_url="/docs",
    redoc_url="/redoc",
    lifespan=lifespan,
)

# Add CORS middleware to allow frontend connections
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Allow all origins for testing
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Simple data models
class Transaction(BaseModel):
    id: str
    amount: float
    status: str
    timestamp: str

class Customer(BaseModel):
    id: str
    name: str
    risk_score: float
    status: str

class Alert(BaseModel):
    id: str
    type: str
    severity: str
    message: str
    timestamp: str

# Mock data storage
mock_transactions: List[Transaction] = []
mock_customers: List[Customer] = []
mock_alerts: List[Alert] = []

# Initialize some mock data
def init_mock_data():
    global mock_transactions, mock_customers, mock_alerts
    
    # Mock transactions
    for i in range(5):
        mock_transactions.append(Transaction(
            id=f"TXN{1000+i}",
            amount=round(random.uniform(100, 10000), 2),
            status=random.choice(["completed", "pending", "flagged"]),
            timestamp=datetime.now().isoformat()
        ))
    
    # Mock customers
    for i in range(5):
        mock_customers.append(Customer(
            id=f"CUST{100+i}",
            name=f"Customer {i+1}",
            risk_score=round(random.uniform(0, 100), 2),
            status=random.choice(["active", "suspicious", "blocked"])
        ))
    
    # Mock alerts
    for i in range(3):
        mock_alerts.append(Alert(
            id=f"ALERT{200+i}",
            type=random.choice(["suspicious_transaction", "unusual_pattern", "high_risk"]),
            severity=random.choice(["low", "medium", "high"]),
            message=f"Alert message {i+1}",
            timestamp=datetime.now().isoformat()
        ))

# Initialize mock data on startup
init_mock_data()


# Root endpoint
@app.get("/")
async def root():
    """Root endpoint."""
    return {
        "message": "Fraud Detection Backend API (Test Mode)",
        "version": "1.0.0-test",
        "status": "running",
        "docs": "/docs",
        "health": "/health"
    }


# Health check endpoint
@app.get("/health")
async def health_check():
    """Health check endpoint."""
    return {
        "status": "healthy",
        "mode": "test",
        "timestamp": datetime.now().isoformat(),
        "components": {
            "api": "healthy",
            "database": "mock"
        }
    }


# ----- Auth (LDAP/AD + JWT 8h) -----
class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: Dict[str, Any]


def _authenticate_demo(username: str, password: str) -> Optional[Dict[str, Any]]:
    """If demo login is enabled and credentials match, return user info. Otherwise None."""
    s = get_settings()
    if not s.demo_login_enabled:
        return None
    if username.strip().lower() == s.demo_username.lower() and password == s.demo_password:
        return {
            "id": username.strip(),
            "name": "Demo User",
            "email": f"{username.strip()}@local",
            "role": "viewer",
            "permissions": ["view"],
        }
    return None


def _authenticate_ldap(username: str, password: str) -> Dict[str, Any]:
    """Validate username/password against AD. Returns user info dict or raises ValueError."""
    s = get_settings()
    if not s.ad_server or not s.ad_search_base:
        demo = _authenticate_demo(username, password)
        if demo is not None:
            return demo
        raise ValueError("AD not configured (AD_SERVER, AD_SEARCH_BASE). Set DEMO_LOGIN_ENABLED=true to use demo login.")
    try:
        from ldap3 import Server, Connection, ALL, SUBTREE
        from ldap3.core.exceptions import LDAPException
    except ImportError:
        raise ValueError("ldap3 not installed; pip install ldap3")

    server = Server(
        s.ad_server,
        port=s.ad_port,
        use_ssl=s.ad_use_ssl,
        get_info=ALL,
    )

    # Bind as service account: try DOMAIN\user then user@domain (some ADs require UPN)
    bind_candidates = [s.ad_user]
    if s.ad_domain and "\\" not in s.ad_user and "@" not in s.ad_user:
        bind_candidates = [
            f"{s.ad_domain}\\{s.ad_user}",
            f"{s.ad_user}@{s.ad_domain}",
        ]
    conn = None
    for bind_user in bind_candidates:
        conn = Connection(
            server,
            user=bind_user,
            password=s.ad_user_password,
            auto_bind=False,
            receive_timeout=s.ad_timeout,
        )
        if conn.bind():
            break
        conn.unbind()
        conn = None

    if conn is None or not conn.bound:
        # Fallback: authenticate user directly (no service account needed)
        return _authenticate_ldap_direct_bind(server, username, password, s.ad_domain, s.ad_timeout)

    try:
        # Search for user by sAMAccountName or userPrincipalName
        conn.search(
            search_base=s.ad_search_base,
            search_filter=f"(|(sAMAccountName={username})(userPrincipalName={username})(userPrincipalName={username}@{s.ad_domain}))",
            search_scope=SUBTREE,
            attributes=["cn", "mail", "sAMAccountName", "userPrincipalName"],
        )
        if not conn.entries:
            raise ValueError("Invalid username or password")
        entry = conn.entries[0]
        user_dn = str(entry.entry_dn)
        mail = getattr(entry, "mail", None) and str(entry.mail) or f"{username}@{s.ad_domain}"
        display_name = getattr(entry, "cn", None) and str(entry.cn) or username
    finally:
        conn.unbind()

    # Verify password by binding as the user
    user_conn = Connection(
        server,
        user=user_dn,
        password=password,
        auto_bind=False,
        receive_timeout=s.ad_timeout,
    )
    try:
        if not user_conn.bind():
            raise ValueError("Invalid username or password")
    finally:
        try:
            user_conn.unbind()
        except Exception:
            pass

    return {
        "id": username,
        "name": display_name,
        "email": mail,
        "role": "viewer",
        "permissions": ["view"],
    }


def _authenticate_ldap_direct_bind(server, username: str, password: str, ad_domain: str, timeout: int) -> Dict[str, Any]:
    """Authenticate by binding directly as the user (no service account)."""
    from ldap3 import Connection
    # Try DOMAIN\username then username@domain (skip if no domain)
    bind_candidates = []
    if ad_domain:
        bind_candidates = [f"{ad_domain}\\{username}", f"{username}@{ad_domain}"]
    else:
        bind_candidates = [username]
    for bind_user in bind_candidates:
        conn = Connection(
            server,
            user=bind_user,
            password=password,
            auto_bind=False,
            receive_timeout=timeout,
        )
        if conn.bind():
            try:
                conn.unbind()
            except Exception:
                pass
            return {
                "id": username,
                "name": username,
                "email": f"{username}@{ad_domain}" if ad_domain else f"{username}@local",
                "role": "viewer",
                "permissions": ["view"],
            }
        try:
            conn.unbind()
        except Exception:
            pass
    raise ValueError("Invalid username or password")


def _is_connection_error(exc: Exception) -> bool:
    """True if this is a network/AD unreachable error (timeout, connection refused, etc.)."""
    msg = str(exc).lower()
    if "10060" in msg or "timed out" in msg or "timeout" in msg or "connection" in msg:
        return True
    if isinstance(exc, (OSError, ConnectionError)):
        return True
    return False


@app.post("/api/v1/auth/login", response_model=LoginResponse)
async def login(payload: LoginRequest):
    """Authenticate via AD/LDAP and return JWT (8h expiry). Demo login supported when enabled."""
    if not payload.username.strip() or not payload.password:
        raise HTTPException(status_code=400, detail="Username and password required")
    try:
        user_info = await asyncio.to_thread(
            _authenticate_ldap,
            payload.username.strip(),
            payload.password,
        )
    except ValueError as e:
        raise HTTPException(status_code=401, detail=str(e))
    except Exception as e:
        s = get_settings()
        if _is_connection_error(e) and s.demo_login_enabled:
            demo = _authenticate_demo(payload.username.strip(), payload.password)
            if demo is not None:
                user_info = demo
            else:
                raise HTTPException(
                    status_code=503,
                    detail="Active Directory is unreachable (connection timed out). Check VPN and AD_SERVER, or use demo login.",
                )
        elif _is_connection_error(e):
            raise HTTPException(
                status_code=503,
                detail="Active Directory is unreachable (connection timed out). Check VPN, network, and AD_SERVER. To log in without AD, set DEMO_LOGIN_ENABLED=true and DEMO_USERNAME/DEMO_PASSWORD in .env.",
            )
        else:
            raise HTTPException(status_code=502, detail=f"Authentication failed: {str(e)}")

    # Allowlist: only usernames in SQLite can log in (when enabled)
    s = get_settings()
    if s.access_allowlist_enabled:
        if not access_allowlist.is_allowed(user_info["id"]):
            raise HTTPException(
                status_code=403,
                detail="You are not allowed to access this application. Contact an administrator to be added to the allowlist.",
            )

    expire = datetime.utcnow() + timedelta(minutes=s.jwt_access_token_expire_minutes)
    to_encode = {
        "sub": user_info["id"],
        "name": user_info["name"],
        "email": user_info["email"],
        "role": user_info["role"],
        "permissions": user_info.get("permissions", []),
        "exp": expire,
        "iat": datetime.utcnow(),
    }
    token = jwt.encode(to_encode, s.jwt_secret, algorithm=s.jwt_algorithm)
    return LoginResponse(access_token=token, user=user_info)


# ----- Access allowlist (SQLite): list / add / remove usernames allowed to log in -----
def _require_access_admin(x_admin_secret: Optional[str] = Header(None, alias="X-Admin-Secret")):
    s = get_settings()
    if not s.access_list_admin_secret:
        raise HTTPException(
            status_code=503,
            detail="Allowlist admin not configured. Set ACCESS_LIST_ADMIN_SECRET in .env",
        )
    if x_admin_secret != s.access_list_admin_secret:
        raise HTTPException(status_code=403, detail="Invalid or missing X-Admin-Secret header")
    return True


@app.get("/api/v1/access/allowed")
async def list_allowed_users(_: bool = Depends(_require_access_admin)):
    """List all usernames that are allowed to log in. Requires X-Admin-Secret header."""
    return {"success": True, "data": access_allowlist.list_allowed()}


class AddAllowedRequest(BaseModel):
    username: str


@app.post("/api/v1/access/allowed")
async def add_allowed_user(
    payload: AddAllowedRequest,
    _: bool = Depends(_require_access_admin),
):
    """Add a username to the allowlist. Requires X-Admin-Secret header."""
    username = (payload.username or "").strip()
    if not username:
        raise HTTPException(status_code=400, detail="username is required")
    added = access_allowlist.add_allowed(username)
    return {"success": True, "added": added, "username": username}


@app.post("/api/v1/access/allowed/reload-from-file")
async def reload_allowed_from_file(_: bool = Depends(_require_access_admin)):
    """Reload allowed usernames from app/data/allowed_usernames.txt (one per line). Requires X-Admin-Secret header."""
    processed, added = access_allowlist.sync_from_file()
    return {"success": True, "processed": processed, "added": added}


@app.delete("/api/v1/access/allowed/{username}")
async def remove_allowed_user(
    username: str,
    _: bool = Depends(_require_access_admin),
):
    """Remove a username from the allowlist. Requires X-Admin-Secret header."""
    removed = access_allowlist.remove_allowed(username)
    return {"success": True, "removed": removed, "username": username.strip()}


# Metrics endpoints
@app.get("/api/v1/metrics")
async def get_metrics():
    """Get system metrics."""
    return {
        "total_transactions": len(mock_transactions),
        "total_customers": len(mock_customers),
        "total_alerts": len(mock_alerts),
        "flagged_transactions": len([t for t in mock_transactions if t.status == "flagged"]),
        "high_risk_customers": len([c for c in mock_customers if c.risk_score > 70]),
        "timestamp": datetime.now().isoformat()
    }


# Transaction endpoints
@app.get("/api/v1/transactions")
async def get_transactions(limit: int = 10):
    """Get all transactions."""
    return {
        "success": True,
        "data": mock_transactions[:limit],
        "total": len(mock_transactions)
    }


@app.get("/api/v1/transactions/{transaction_id}")
async def get_transaction(transaction_id: str):
    """Get a specific transaction."""
    transaction = next((t for t in mock_transactions if t.id == transaction_id), None)
    if not transaction:
        raise HTTPException(status_code=404, detail="Transaction not found")
    return {
        "success": True,
        "data": transaction
    }


# Customer endpoints
@app.get("/api/v1/customers")
async def get_customers(limit: int = 10):
    """Get all customers."""
    return {
        "success": True,
        "data": mock_customers[:limit],
        "total": len(mock_customers)
    }


@app.get("/api/v1/customers/{customer_id}")
async def get_customer(customer_id: str):
    """Get a specific customer."""
    customer = next((c for c in mock_customers if c.id == customer_id), None)
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return {
        "success": True,
        "data": customer
    }


# Alert endpoints
@app.get("/api/v1/alerts")
async def get_alerts(limit: int = 10):
    """Get all alerts."""
    return {
        "success": True,
        "data": mock_alerts[:limit],
        "total": len(mock_alerts)
    }


@app.get("/api/v1/alerts/{alert_id}")
async def get_alert(alert_id: str):
    """Get a specific alert."""
    alert = next((a for a in mock_alerts if a.id == alert_id), None)
    if not alert:
        raise HTTPException(status_code=404, detail="Alert not found")
    return {
        "success": True,
        "data": alert
    }


# Test data generation endpoints
@app.post("/api/v1/test/generate-transaction")
async def generate_transaction():
    """Generate a new mock transaction."""
    new_transaction = Transaction(
        id=f"TXN{1000 + len(mock_transactions)}",
        amount=round(random.uniform(100, 10000), 2),
        status=random.choice(["completed", "pending", "flagged"]),
        timestamp=datetime.now().isoformat()
    )
    mock_transactions.append(new_transaction)
    return {
        "success": True,
        "message": "Transaction generated",
        "data": new_transaction
    }


@app.post("/api/v1/test/generate-alert")
async def generate_alert():
    """Generate a new mock alert."""
    new_alert = Alert(
        id=f"ALERT{200 + len(mock_alerts)}",
        type=random.choice(["suspicious_transaction", "unusual_pattern", "high_risk"]),
        severity=random.choice(["low", "medium", "high"]),
        message=f"Test alert generated at {datetime.now().strftime('%H:%M:%S')}",
        timestamp=datetime.now().isoformat()
    )
    mock_alerts.append(new_alert)
    return {
        "success": True,
        "message": "Alert generated",
        "data": new_alert
    }


@app.post("/api/v1/test/reset-data")
async def reset_data():
    """Reset all mock data."""
    init_mock_data()
    return {
        "success": True,
        "message": "All data has been reset",
        "timestamp": datetime.now().isoformat()
    }


# Account endpoints (simplified)
@app.get("/api/v1/accounts")
async def get_accounts():
    """Get mock accounts."""
    return {
        "success": True,
        "data": [
            {"id": "ACC001", "balance": 10000.00, "status": "active"},
            {"id": "ACC002", "balance": 25000.00, "status": "active"},
            {"id": "ACC003", "balance": 5000.00, "status": "flagged"}
        ],
        "total": 3
    }


# Compliance Register endpoints
# Keep compliance data path relative to this repository so it works on any machine.
DATA_DIR = Path(__file__).resolve().parent / "data"
COMPLIANCE_FILE = DATA_DIR / "Compliance_Register - Copy.xlsx"
EMAIL_FILE = DATA_DIR / "email.xlsx"
if not EMAIL_FILE.exists():
    EMAIL_FILE = DATA_DIR / "email"  # try no extension

# Simple cache for combined dataframe (cleared on updates)
_cached_df = None
_cached_sheet_map = None
_cache_timestamp = None


def _forward_fill_law_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Forward-fill LAW/REGULATION/DIRECTIVE so one regulation applies to every row it references (handles '-' and merged cells)."""
    if df.empty:
        return df
    law_cols = [c for c in df.columns if isinstance(c, str) and "law" in c.lower() and ("regulation" in c.lower() or "directive" in c.lower())]
    if not law_cols:
        law_cols = [c for c in df.columns if isinstance(c, str) and "law" in c.lower()]
    for col in law_cols:
        try:
            # Treat empty, '-', and NaN (merged cells in Excel) as "same as previous"
            s = df[col].astype(str).str.strip()
            s = s.replace(["", "-", "nan", "NaN", "None"], pd.NA)
            df[col] = s.ffill().bfill()
        except Exception:
            pass
    return df


def read_all_sheets_from_excel(excel_path: Path, use_cache: bool = True):
    """Read all sheets from Excel file and combine them into one DataFrame."""
    global _cached_df, _cached_sheet_map, _cache_timestamp
    
    # Use cache if available and file hasn't been modified
    if use_cache and _cached_df is not None and _cached_sheet_map is not None:
        if excel_path.exists():
            file_mtime = excel_path.stat().st_mtime
            if _cache_timestamp is not None and file_mtime <= _cache_timestamp:
                return _cached_df.copy(), _cached_sheet_map.copy()
    
    try:
        # Suppress openpyxl "Data Validation extension is not supported" warning
        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", message=".*Data Validation extension.*", category=UserWarning)
            excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
        sheet_names = excel_file.sheet_names

        all_dataframes = []
        sheet_data_map = {}  # Store original sheet data for writing back

        non_empty_count = 0
        for sheet_name in sheet_names:
            try:
                with warnings.catch_warnings():
                    warnings.filterwarnings("ignore", message=".*Data Validation extension.*", category=UserWarning)
                    df = pd.read_excel(excel_path, sheet_name=sheet_name, engine='openpyxl')
                
                # Skip empty sheets
                if df.empty:
                    sheet_data_map[sheet_name] = df
                    continue
                
                # Load all non-empty sheets (departments = sheet names)
                non_empty_count += 1

                # Store original sheet data
                sheet_data_map[sheet_name] = df.copy()
                
                # Add a column to track the source sheet (department name)
                # Only add if it doesn't already exist
                if 'SOURCE_SHEET' not in df.columns:
                    df['SOURCE_SHEET'] = sheet_name
                
                # Also try to set department column if it doesn't exist or is empty
                dept_col = None
                for col in df.columns:
                    if 'department' in col.lower() or 'responsible' in col.lower():
                        dept_col = col
                        break
                
                # If department column exists but has empty values, fill with sheet name
                if dept_col:
                    df[dept_col] = df[dept_col].fillna(sheet_name)
                elif 'RESPONSIBLE DEPARTMENT' not in df.columns and 'Responsible Department' not in df.columns:
                    # Add department column if it doesn't exist
                    df['RESPONSIBLE DEPARTMENT'] = sheet_name
                
                all_dataframes.append(df)
            except Exception as e:
                # Skip sheets that can't be read
                print(f"Warning: Could not read sheet '{sheet_name}': {str(e)}")
                sheet_data_map[sheet_name] = pd.DataFrame()
                continue
        
        if not all_dataframes:
            return pd.DataFrame(), sheet_data_map
        
        # Combine all dataframes
        combined_df = pd.concat(all_dataframes, ignore_index=True)
        
        # Forward-fill LAW/REGULATION/DIRECTIVE (and similar) when cell is "-" or empty so the law spans all references
        combined_df = _forward_fill_law_columns(combined_df)
        
        # Cache the result
        _cached_df = combined_df.copy()
        _cached_sheet_map = sheet_data_map.copy()
        _cache_timestamp = excel_path.stat().st_mtime if excel_path.exists() else None
        
        return combined_df, sheet_data_map
        
    except Exception as e:
        import traceback
        error_detail = f"Error reading Excel sheets: {str(e)}\nTraceback: {traceback.format_exc()}"
        raise Exception(error_detail)

def save_to_all_sheets(excel_path: Path, combined_df: pd.DataFrame, sheet_data_map: Dict[str, pd.DataFrame]):
    """Save combined dataframe back to individual sheets based on SOURCE_SHEET column."""
    try:
        from openpyxl import load_workbook
        
        # Load the workbook
        wb = load_workbook(excel_path)
        
        # Get all sheet names from the workbook
        existing_sheets = wb.sheetnames.copy()
        
        # Group records by source sheet
        if 'SOURCE_SHEET' in combined_df.columns:
            # Get unique sheet names from the combined dataframe
            unique_sheets = combined_df['SOURCE_SHEET'].unique().tolist()
            
            # Process each sheet
            for sheet_name in unique_sheets:
                # Get records for this sheet
                sheet_df = combined_df[combined_df['SOURCE_SHEET'] == sheet_name].copy()
                
                # Drop SOURCE_SHEET and 'id' columns before saving
                sheet_df = sheet_df.drop(columns=['SOURCE_SHEET', 'id'], errors='ignore')
                
                # Remove the sheet if it exists
                if sheet_name in wb.sheetnames:
                    wb.remove(wb[sheet_name])
                
                # Create new sheet
                ws = wb.create_sheet(sheet_name)
                
                # Write headers
                if not sheet_df.empty:
                    headers = list(sheet_df.columns)
                    ws.append(headers)
                    
                    # Write data
                    for _, row in sheet_df.iterrows():
                        row_data = []
                        for col in headers:
                            val = row[col]
                            if pd.isna(val):
                                row_data.append(None)
                            else:
                                row_data.append(val)
                        ws.append(row_data)
                else:
                    # Create empty sheet with headers from original
                    if sheet_name in sheet_data_map and not sheet_data_map[sheet_name].empty:
                        headers = list(sheet_data_map[sheet_name].columns)
                        ws.append(headers)
        
        # Save the workbook
        wb.save(excel_path)
        
        # Clear cache after saving
        global _cached_df, _cached_sheet_map, _cache_timestamp
        _cached_df = None
        _cached_sheet_map = None
        _cache_timestamp = None
        
    except Exception as e:
        import traceback
        print(f"Error saving to sheets: {str(e)}\nTraceback: {traceback.format_exc()}")
        # Fallback: save to first sheet only
        try:
            df_to_save = combined_df.drop(columns=['SOURCE_SHEET', 'id'], errors='ignore')
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                df_to_save.to_excel(writer, index=False, sheet_name='Sheet1')
        except Exception as e2:
            raise Exception(f"Failed to save Excel file: {str(e2)}")

@app.get("/api/v1/compliance/cache-status")
async def get_compliance_cache_status():
    """Check if Redis and SQLite cache are active. Use to verify why loading is slow (e.g. redis_connected: false)."""
    s = get_settings()
    redis_ok = redis_cache.is_connected()
    return {
        "redis_connected": redis_ok,
        "use_redis_cache": getattr(s, "use_redis_cache", False),
        "use_compliance_sqlite": getattr(s, "use_compliance_sqlite", False),
        "sqlite_has_data": compliance_db.has_data() if getattr(s, "use_compliance_sqlite", False) else False,
        "hint": "If redis_connected is false, start Redis (e.g. docker run -p 6379:6379 redis:7-alpine) and set USE_REDIS_CACHE=true.",
    }


@app.get("/api/v1/compliance/debug-path")
async def get_compliance_debug_path():
    """Debug endpoint to check the file path being used."""
    return {
        "data_dir": str(DATA_DIR),
        "compliance_file": str(COMPLIANCE_FILE),
        "file_exists": COMPLIANCE_FILE.exists(),
        "data_dir_exists": DATA_DIR.exists()
    }


def _read_owners_from_excel() -> tuple:
    """Read owners and email from email Excel. Returns (list, source, message)."""
    fallback = [
        {"owner": "Jane Doe", "email": "jane.doe@example.com"},
        {"owner": "John Smith", "email": "john.smith@example.com"},
        {"owner": "Alice Brown", "email": "alice.brown@example.com"},
        {"owner": "Bob Wilson", "email": "bob.wilson@example.com"},
        {"owner": "Carol Davis", "email": "carol.davis@example.com"},
    ]
    path = DATA_DIR / "email.xlsx"
    if not path.exists():
        path = DATA_DIR / "email"
    if not path.exists():
        return fallback, "fallback", "Email file not found (expected email.xlsx in app/data). Using default list."
    try:
        df = pd.read_excel(path, sheet_name=0, engine="openpyxl")
        if df.empty or len(df.columns) < 2:
            return fallback, "fallback", "Email file is empty or has no columns. Using default list."
        cols = [c for c in df.columns if isinstance(c, str)]
        owner_col = next((c for c in cols if "owner" in c.lower()), None)
        email_col = next((c for c in cols if "email" in c.lower()), None)
        if not owner_col or not email_col:
            return fallback, "fallback", "Email file must have columns named 'owners' and 'email'. Using default list."
        out = []
        for _, row in df.iterrows():
            o, e = row.get(owner_col), row.get(email_col)
            if pd.isna(o) and pd.isna(e):
                continue
            out.append({
                "owner": str(o).strip() if not pd.isna(o) else "",
                "email": str(e).strip() if not pd.isna(e) else "",
            })
        if not out:
            return fallback, "fallback", "Email file has no data rows. Using default list."
        return out, "file", f"Loaded {len(out)} owner(s) from email.xlsx"
    except Exception as e:
        return fallback, "fallback", f"Could not read email file: {e!s}. Using default list."


def _timing_headers(redis_ms: float, excel_ms: float, build_ms: float, redis_set_ms: float, total_ms: float, source: str) -> dict:
    """Build response headers for latency debugging."""
    return {
        "X-Timing-Redis-Ms": f"{redis_ms:.1f}",
        "X-Timing-Excel-Ms": f"{excel_ms:.1f}",
        "X-Timing-Build-Ms": f"{build_ms:.1f}",
        "X-Timing-RedisSet-Ms": f"{redis_set_ms:.1f}",
        "X-Timing-Total-Ms": f"{total_ms:.1f}",
        "X-Timing-Source": source,
    }


@app.get("/api/v1/compliance/initial")
async def get_compliance_initial(
    limit: int = 1000,
    offset: int = 0,
    department: Optional[str] = None,
    status: Optional[str] = None,
):
    """Single-call bootstrap: departments + statuses + first page. Response headers show latency breakdown (X-Timing-*)."""
    t0 = time.perf_counter()
    cache_key = f"compliance:initial:{limit}:{offset}:{(department or '').strip()}:{(status or '').strip()}"
    cached = redis_cache.cache_get(cache_key)
    t1 = time.perf_counter()
    redis_ms = (t1 - t0) * 1000
    if cached is not None:
        total_ms = (t1 - t0) * 1000
        return JSONResponse(
            content=cached,
            headers=_timing_headers(redis_ms, 0, 0, 0, total_ms, "redis"),
        )
    # Prefer SQLite (fast) when enabled and populated
    s = get_settings()
    if s.use_compliance_sqlite and compliance_db.has_data():
        t_sqlite_start = time.perf_counter()
        departments = compliance_db.get_departments()
        statuses = compliance_db.get_statuses()
        records, total = compliance_db.get_records(limit, offset, department, status)
        for rec in records:
            for k, v in list(rec.items()):
                if v is not None and isinstance(v, float) and (v != v or v == 1e999):
                    rec[k] = None
        offset_safe = max(0, min(offset, total))
        t_sqlite_end = time.perf_counter()
        out = {
            "success": True,
            "departments": departments,
            "statuses": statuses,
            "records": {"data": records, "total": total, "limit": limit, "offset": offset_safe},
        }
        redis_cache.cache_set(cache_key, out, s.redis_cache_ttl)
        total_ms = (time.perf_counter() - t0) * 1000
        sqlite_ms = (t_sqlite_end - t_sqlite_start) * 1000
        return JSONResponse(
            content=out,
            headers={
                "X-Timing-Redis-Ms": f"{redis_ms:.1f}",
                "X-Timing-Excel-Ms": "0",
                "X-Timing-Build-Ms": f"{sqlite_ms:.1f}",
                "X-Timing-RedisSet-Ms": "0",
                "X-Timing-Total-Ms": f"{total_ms:.1f}",
                "X-Timing-Source": "sqlite",
            },
        )
    excel_path = COMPLIANCE_FILE
    if not excel_path.exists():
        out = {
            "success": True,
            "departments": [],
            "statuses": [],
            "records": {"data": [], "total": 0, "limit": limit, "offset": 0},
        }
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        t_end = time.perf_counter()
        return JSONResponse(
            content=out,
            headers=_timing_headers(redis_ms, 0, 0, (t_end - t1) * 1000, (t_end - t0) * 1000, "excel-missing"),
        )
    try:
        df, _ = read_all_sheets_from_excel(excel_path)
    except Exception:
        out = {
            "success": False,
            "departments": [],
            "statuses": [],
            "records": {"data": [], "total": 0, "limit": limit, "offset": 0},
        }
        t_end = time.perf_counter()
        return JSONResponse(content=out, headers=_timing_headers(redis_ms, (t_end - t1) * 1000, 0, 0, (t_end - t0) * 1000, "excel-error"))
    t2 = time.perf_counter()
    excel_ms = (t2 - t1) * 1000
    if df.empty:
        out = {
            "success": True,
            "departments": [],
            "statuses": [],
            "records": {"data": [], "total": 0, "limit": limit, "offset": 0},
        }
        t_before_set = time.perf_counter()
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        t_end = time.perf_counter()
        return JSONResponse(
            content=out,
            headers=_timing_headers(redis_ms, excel_ms, (t_before_set - t2) * 1000, (t_end - t_before_set) * 1000, (t_end - t0) * 1000, "excel"),
        )
    df = df.copy()
    df["id"] = df.index
    if "SOURCE_SHEET" in df.columns:
        depts = df["SOURCE_SHEET"].dropna().astype(str).str.strip().unique().tolist()
        departments = sorted(set(depts))
    else:
        try:
            excel_file = pd.ExcelFile(excel_path, engine="openpyxl")
            departments = sorted([str(s) for s in excel_file.sheet_names if s and str(s).strip()])
        except Exception:
            departments = []
    status_col = None
    for col in df.columns:
        if "comment" in col.lower():
            continue
        if "status" in col.lower() or "compliance" in col.lower():
            status_col = col
            break
    statuses = sorted([v for v in (df[status_col].dropna().astype(str).str.strip().unique().tolist() if status_col else []) if v])
    if department and "SOURCE_SHEET" in df.columns:
        try:
            df = df[df["SOURCE_SHEET"].astype(str).str.strip().str.lower() == department.strip().lower()]
        except Exception:
            pass
    elif department:
        dept_col = next((c for c in df.columns if "department" in c.lower() or "responsible" in c.lower()), None)
        if dept_col:
            try:
                df = df[df[dept_col].astype(str).str.contains(department, case=False, na=False)]
            except Exception:
                pass
    if status and status_col:
        try:
            df = df[df[status_col].astype(str).str.contains(status, case=False, na=False)]
        except Exception:
            pass
    df = df.reset_index(drop=True)
    total = len(df)
    offset_safe = max(0, min(offset, total))
    end_idx = min(offset_safe + limit, total)
    records = df.iloc[offset_safe:end_idx].to_dict("records")
    for record in records:
        for key, value in record.items():
            if pd.isna(value):
                record[key] = None
    t3 = time.perf_counter()
    build_ms = (t3 - t2) * 1000
    out = {
        "success": True,
        "departments": departments,
        "statuses": statuses,
        "records": {"data": records, "total": total, "limit": limit, "offset": offset_safe},
    }
    redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
    t4 = time.perf_counter()
    redis_set_ms = (t4 - t3) * 1000
    total_ms = (t4 - t0) * 1000
    return JSONResponse(
        content=out,
        headers=_timing_headers(redis_ms, excel_ms, build_ms, redis_set_ms, total_ms, "excel"),
    )


@app.get("/api/v1/compliance/departments")
async def get_compliance_departments():
    """Return all department names = sheet names. Uses Redis cache when enabled, then in-memory cache."""
    cache_key = "compliance:departments"
    cached = redis_cache.cache_get(cache_key)
    if cached is not None:
        return cached
    try:
        global _cached_df
        excel_path = COMPLIANCE_FILE
        if not excel_path.exists():
            out = {"success": True, "data": [], "message": "Compliance file not found."}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        if _cached_df is None:
            try:
                _cached_df, _ = read_all_sheets_from_excel(excel_path)
            except Exception:
                pass
        if _cached_df is not None and not _cached_df.empty and "SOURCE_SHEET" in _cached_df.columns:
            depts = _cached_df["SOURCE_SHEET"].dropna().astype(str).str.strip().unique().tolist()
            out = {"success": True, "data": sorted(set(depts)), "message": "From cache."}
        else:
            excel_file = pd.ExcelFile(excel_path, engine="openpyxl")
            sheet_names = [str(s) for s in excel_file.sheet_names if s and str(s).strip()]
            out = {"success": True, "data": sorted(sheet_names), "message": "Sheet names."}
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        return out
    except Exception as e:
        return {"success": False, "data": [], "message": str(e)}


@app.get("/api/v1/compliance/statuses")
async def get_compliance_statuses():
    """Return unique compliance status values. Uses Redis cache when enabled."""
    cache_key = "compliance:statuses"
    cached = redis_cache.cache_get(cache_key)
    if cached is not None:
        return cached
    try:
        excel_path = COMPLIANCE_FILE
        if not excel_path.exists():
            out = {"success": True, "data": []}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        df, _ = read_all_sheets_from_excel(excel_path)
        if df.empty:
            out = {"success": True, "data": []}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        status_col = None
        for col in df.columns:
            if "comment" in col.lower():
                continue
            if "status" in col.lower() or "compliance" in col.lower():
                status_col = col
                break
        if not status_col:
            out = {"success": True, "data": []}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        values = df[status_col].dropna().astype(str).str.strip().unique()
        values = sorted([v for v in values if v])
        out = {"success": True, "data": values}
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        return out
    except Exception as e:
        return {"success": False, "data": [], "message": str(e)}


@app.get("/api/v1/compliance/owners")
async def get_compliance_owners():
    """Return list of { owner, email } from email Excel (owners + email columns)."""
    data, source, message = _read_owners_from_excel()
    return {"success": True, "data": data, "source": source, "message": message}


class SendEmailRequest(BaseModel):
    to_email: str
    subject: Optional[str] = None
    body: Optional[str] = None
    automated: bool = False


class NotifyOwnersRequest(BaseModel):
    to_emails: List[str]
    subject: str
    body: Optional[str] = None


def _send_email_sync(to_email: str, subject: str, body: str) -> None:
    """Send a single email via SMTP (blocking). Raises on failure."""
    s = get_settings()
    if not s.mail_username or not s.mail_password:
        raise RuntimeError("SMTP not configured: set MAIL_USERNAME and MAIL_PASSWORD in .env")
    sender = s.mail_default_sender or s.mail_username
    msg = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["From"] = sender
    msg["To"] = to_email
    msg.attach(MIMEText(body or "", "plain", "utf-8"))
    with smtplib.SMTP(s.mail_server, s.mail_port) as server:
        if s.mail_use_tls:
            server.starttls()
        server.login(s.mail_username, s.mail_password)
        server.sendmail(sender, to_email, msg.as_string())


@app.post("/api/v1/compliance/send-email")
async def send_compliance_email(payload: SendEmailRequest):
    """Send email via SMTP (Office 365). Subject required for customized; automated uses default subject/body."""
    if not payload.automated and not payload.subject:
        raise HTTPException(status_code=400, detail="Subject is required for customized email.")
    subject = payload.subject or "Compliance notification"
    body = payload.body or "This is an automated compliance notification from the RegMgmt system."
    try:
        await asyncio.to_thread(_send_email_sync, payload.to_email, subject, body)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Failed to send email: {str(e)}")
    return {
        "success": True,
        "message": "Email sent successfully.",
        "to": payload.to_email,
        "automated": payload.automated,
    }


@app.post("/api/v1/compliance/notify-owners")
async def notify_compliance_owners(payload: NotifyOwnersRequest):
    """Send the same notification to multiple owner emails (e.g. all owners of currently filtered records)."""
    if not payload.to_emails:
        raise HTTPException(status_code=400, detail="to_emails is required and cannot be empty.")
    subject = (payload.subject or "").strip() or "Compliance notification"
    body = (payload.body or "").strip() or "This is an automated compliance notification from the RegMgmt system."
    seen = set()
    unique_emails = []
    for e in payload.to_emails:
        e = (e or "").strip()
        if e and e not in seen:
            seen.add(e)
            unique_emails.append(e)
    sent, failed = 0, 0
    errors = []
    for to_email in unique_emails:
        try:
            await asyncio.to_thread(_send_email_sync, to_email, subject, body)
            sent += 1
        except Exception as e:
            failed += 1
            errors.append(f"{to_email}: {str(e)}")
    return {
        "success": failed == 0,
        "message": f"Sent to {sent} recipient(s)." + (f" Failed: {failed}." if failed else ""),
        "sent": sent,
        "failed": failed,
        "errors": errors if errors else None,
    }


@app.post("/api/v1/compliance/sync-from-excel")
async def sync_compliance_from_excel():
    """Re-import the Compliance Register Excel file into SQLite. Call after updating the Excel file. Broadcasts to /ws clients."""
    if not COMPLIANCE_FILE.exists():
        raise HTTPException(status_code=404, detail=f"Compliance file not found: {COMPLIANCE_FILE}")
    try:
        count = await asyncio.to_thread(
            compliance_db.import_from_excel,
            read_all_sheets_from_excel,
            COMPLIANCE_FILE,
        )
        return {"success": True, "message": f"Imported {count} records into SQLite.", "count": count}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/v1/compliance/stats")
async def get_compliance_stats():
    """Get statistics from the Compliance Register Excel file."""
    try:
        # Use the absolute path to the Excel file
        excel_path = COMPLIANCE_FILE
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Compliance Register file not found at: {excel_path}. Please ensure the file exists at this location."
            )
        
        # Read all sheets from the Excel file
        try:
            df, _ = read_all_sheets_from_excel(excel_path)
        except ImportError as e:
            raise HTTPException(
                status_code=500, 
                detail=f"openpyxl library is required to read Excel files. Please install it with: pip install openpyxl. Error: {str(e)}"
            )
        except FileNotFoundError:
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found at: {excel_path}"
            )
        except Exception as e:
            import traceback
            error_detail = f"Error reading Excel file: {str(e)}\nFile path: {excel_path}\nTraceback: {traceback.format_exc()}"
            raise HTTPException(status_code=500, detail=error_detail)
        
        # Check if dataframe is empty
        if df.empty:
            return {
                "success": True,
                "data": {
                    "total_records": 0,
                    "columns": [],
                    "summary": {
                        "total_entries": 0,
                        "data_quality": {
                            "complete_records": 0,
                            "incomplete_records": 0,
                            "completion_rate": "0%"
                        }
                    },
                    "timestamp": datetime.now().isoformat()
                }
            }
        
        # Calculate statistics
        total_records = len(df)
        
        # Get unique values for categorical columns (adjust column names based on your actual Excel file)
        stats = {
            "total_records": total_records,
            "columns": list(df.columns),
            "summary": {
                "total_entries": total_records,
                "data_quality": {
                    "complete_records": int(df.dropna().shape[0]),
                    "incomplete_records": int(df.shape[0] - df.dropna().shape[0]),
                    "completion_rate": f"{(df.dropna().shape[0] / df.shape[0] * 100):.1f}%"
                }
            },
            "timestamp": datetime.now().isoformat()
        }
        
        # Add column-specific statistics
        numeric_cols = df.select_dtypes(include=['number']).columns
        if len(numeric_cols) > 0:
            stats["numeric_summary"] = {}
            for col in numeric_cols[:5]:  # Limit to first 5 numeric columns
                stats["numeric_summary"][col] = {
                    "mean": float(df[col].mean()) if not df[col].isna().all() else 0,
                    "min": float(df[col].min()) if not df[col].isna().all() else 0,
                    "max": float(df[col].max()) if not df[col].isna().all() else 0
                }
        
        return {
            "success": True,
            "data": stats
        }
        
    except HTTPException:
        raise
    except Exception as e:
        import traceback
        error_detail = f"Error reading compliance data: {str(e)}\n{traceback.format_exc()}"
        raise HTTPException(status_code=500, detail=error_detail)


@app.get("/api/v1/compliance/records")
async def get_compliance_records(
    limit: int = 1000, 
    offset: int = 0,
    department: Optional[str] = None,
    status: Optional[str] = None
):
    """Get records from the Compliance Register. Uses SQLite when enabled (fast), else Excel."""
    cache_key = f"compliance:records:{limit}:{offset}:{(department or '').strip()}:{(status or '').strip()}"
    cached = redis_cache.cache_get(cache_key)
    if cached is not None:
        return cached
    if get_settings().use_compliance_sqlite and compliance_db.has_data():
        records, total = compliance_db.get_records(limit, offset, department, status)
        for rec in records:
            for k, v in list(rec.items()):
                if v is not None and isinstance(v, float) and (v != v or v == 1e999):
                    rec[k] = None
        offset_safe = max(0, min(offset, total))
        out = {"success": True, "data": records, "total": total, "limit": limit, "offset": offset_safe}
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        return out
    try:
        excel_path = COMPLIANCE_FILE
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Compliance Register file not found at: {excel_path}. Please ensure the file exists at this location."
            )
        
        # Read all sheets from the Excel file
        try:
            df, _ = read_all_sheets_from_excel(excel_path)
        except ImportError as e:
            raise HTTPException(
                status_code=500, 
                detail=f"openpyxl library is required to read Excel files. Please install it with: pip install openpyxl. Error: {str(e)}"
            )
        except FileNotFoundError:
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found at: {excel_path}"
            )
        except Exception as e:
            import traceback
            error_detail = f"Error reading Excel file: {str(e)}\nFile path: {excel_path}\nTraceback: {traceback.format_exc()}"
            raise HTTPException(status_code=500, detail=error_detail)
        
        # Check if dataframe is empty
        if df.empty:
            out = {"success": True, "data": [], "total": 0, "limit": limit, "offset": offset}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        
        # Work on a copy to avoid fragmentation warning and avoid mutating cached df
        df = df.copy()
        # Store original index as 'id' for API use (overwrite if 'id' column exists)
        df["id"] = df.index
        
        # Apply filters (department = sheet name; filter by SOURCE_SHEET so dropdown matches all sheets)
        if department:
            if "SOURCE_SHEET" in df.columns:
                try:
                    df = df[df["SOURCE_SHEET"].astype(str).str.strip().str.lower() == department.strip().lower()]
                except Exception:
                    pass
            else:
                dept_col = None
                for col in df.columns:
                    if "department" in col.lower() or "responsible" in col.lower():
                        dept_col = col
                        break
                if dept_col:
                    try:
                        df = df[df[dept_col].astype(str).str.contains(department, case=False, na=False)]
                    except Exception:
                        pass
        
        if status:
            # Try to find status column (case-insensitive)
            status_col = None
            for col in df.columns:
                if 'status' in col.lower() or 'compliance' in col.lower():
                    status_col = col
                    break
            if status_col:
                try:
                    df = df[df[status_col].astype(str).str.contains(status, case=False, na=False)]
                except Exception as e:
                    # If filtering fails, continue without filter
                    pass
        
        # Reset index for display, but keep 'id' as original row number
        df = df.reset_index(drop=True)
        
        # Handle empty dataframe
        if len(df) == 0:
            out = {"success": True, "data": [], "total": 0, "limit": limit, "offset": offset}
            redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
            return out
        
        # Ensure offset is within bounds
        offset = max(0, min(offset, len(df)))
        
        # Convert to records
        end_idx = min(offset + limit, len(df))
        records = df.iloc[offset:end_idx].to_dict('records')
        
        # Convert NaN to None for JSON serialization
        for record in records:
            for key, value in record.items():
                if pd.isna(value):
                    record[key] = None
        
        out = {"success": True, "data": records, "total": len(df), "limit": limit, "offset": offset}
        redis_cache.cache_set(cache_key, out, get_settings().redis_cache_ttl)
        return out
        
    except HTTPException:
        raise
    except Exception as e:
        error_detail = f"Error reading compliance records: {str(e)}"
        raise HTTPException(status_code=500, detail=error_detail)


@app.put("/api/v1/compliance/records/{record_id}")
async def update_compliance_record(record_id: int, record_data: Dict[str, Any] = Body(...)):
    """Update a compliance record."""
    try:
        # Use the absolute path to the Excel file
        excel_path = COMPLIANCE_FILE
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Compliance Register file not found at: {excel_path}. Please ensure the file exists at this location."
            )
        
        # Read all sheets from the Excel file
        try:
            df, sheet_data_map = read_all_sheets_from_excel(excel_path)
        except ImportError as e:
            raise HTTPException(
                status_code=500, 
                detail=f"openpyxl library is required to read Excel files. Please install it with: pip install openpyxl. Error: {str(e)}"
            )
        except FileNotFoundError:
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found at: {excel_path}"
            )
        except Exception as e:
            import traceback
            error_detail = f"Error reading Excel file: {str(e)}\nFile path: {excel_path}\nTraceback: {traceback.format_exc()}"
            raise HTTPException(status_code=500, detail=error_detail)
        
        # Check if dataframe is empty
        if df.empty:
            raise HTTPException(status_code=404, detail="No records found in Excel file")
        
        # Store original index as 'id' for API use
        df['id'] = df.index
        
        # Find the record by id (which is the index in our system)
        if record_id < 0 or record_id >= len(df):
            raise HTTPException(status_code=404, detail="Record not found")
        
        # Update the record - exclude 'id', 'index', and 'SOURCE_SHEET' from updates
        for key, value in record_data.items():
            if key in df.columns and key not in ['id', 'index', 'SOURCE_SHEET']:
                df.at[record_id, key] = value
        
        # Save back to all sheets
        try:
            save_to_all_sheets(excel_path, df, sheet_data_map)
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="openpyxl library is required to write Excel files. Please install it with: pip install openpyxl"
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error writing Excel file: {str(e)}")
        
        # Return updated record
        updated_record = df.iloc[record_id].to_dict()
        updated_record['id'] = record_id
        for key, value in updated_record.items():
            if pd.isna(value):
                updated_record[key] = None
        
        return {
            "success": True,
            "message": "Record updated successfully",
            "data": updated_record
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error updating compliance record: {str(e)}")


@app.post("/api/v1/compliance/records")
async def create_compliance_record(record_data: Dict[str, Any] = Body(...)):
    """Add a new compliance record."""
    try:
        # Use the absolute path to the Excel file
        excel_path = COMPLIANCE_FILE
        
        if not excel_path.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"Compliance Register file not found at: {excel_path}. Please ensure the file exists at this location."
            )
        
        # Read all sheets from the Excel file
        try:
            df, sheet_data_map = read_all_sheets_from_excel(excel_path)
        except ImportError as e:
            raise HTTPException(
                status_code=500, 
                detail=f"openpyxl library is required to read Excel files. Please install it with: pip install openpyxl. Error: {str(e)}"
            )
        except FileNotFoundError:
            raise HTTPException(
                status_code=404,
                detail=f"Excel file not found at: {excel_path}"
            )
        except Exception as e:
            import traceback
            error_detail = f"Error reading Excel file: {str(e)}\nFile path: {excel_path}\nTraceback: {traceback.format_exc()}"
            raise HTTPException(status_code=500, detail=error_detail)
        
        # Determine which sheet to add the record to based on department
        target_sheet = None
        dept_col = None
        for col in df.columns:
            if 'department' in col.lower() or 'responsible' in col.lower():
                dept_col = col
                break
        
        if dept_col and dept_col in record_data and record_data[dept_col]:
            target_sheet = str(record_data[dept_col]).strip()
        elif 'SOURCE_SHEET' in df.columns and len(df) > 0:
            # Use the first sheet as default
            target_sheet = df['SOURCE_SHEET'].iloc[0] if 'SOURCE_SHEET' in df.columns else None
        else:
            # Get first sheet name from Excel file
            try:
                excel_file = pd.ExcelFile(excel_path, engine='openpyxl')
                target_sheet = excel_file.sheet_names[0] if excel_file.sheet_names else 'Sheet1'
            except:
                target_sheet = 'Sheet1'
        
        # Create new row - only include columns that exist in the dataframe
        new_row = {}
        for col in df.columns:
            # Skip 'id', 'index', and 'SOURCE_SHEET' columns as they're auto-generated
            if col not in ['id', 'index', 'SOURCE_SHEET']:
                new_row[col] = record_data.get(col, None)
        
        # Set SOURCE_SHEET for the new record
        new_row['SOURCE_SHEET'] = target_sheet
        
        # Append new row
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        
        # Calculate the new record's ID (it will be the last row index)
        new_record_id = len(df) - 1
        df.at[new_record_id, 'id'] = new_record_id
        
        # Save back to all sheets
        try:
            save_to_all_sheets(excel_path, df, sheet_data_map)
        except ImportError:
            raise HTTPException(
                status_code=500, 
                detail="openpyxl library is required to write Excel files. Please install it with: pip install openpyxl"
            )
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error writing Excel file: {str(e)}")
        
        # Return new record with the correct ID
        new_record = df.iloc[-1].to_dict()
        new_record['id'] = new_record_id
        for key, value in new_record.items():
            if pd.isna(value):
                new_record[key] = None
        
        return {
            "success": True,
            "message": "Record created successfully",
            "data": new_record
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error creating compliance record: {str(e)}")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000, reload=False)
